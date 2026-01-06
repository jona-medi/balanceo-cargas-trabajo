# app_streamlit.py
import os
import subprocess
import pandas as pd
import streamlit as st
from datetime import datetime, timedelta, time
import numpy as np
import warnings
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import io
import pytz

warnings.filterwarnings('ignore')
st.set_page_config(page_title="Evaluación - Matriz de Funciones → BALANCE", layout="wide")

# ---------------- CONFIG ----------------
# Obtiene la ruta base del proyecto, funcione donde funcione
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

ROOT_FOLDER = os.path.join(BASE_DIR, "data", "matrices_funciones")  # Ahora apunta a ./data
BALANCE_PATH = os.path.join(BASE_DIR, "data", "BALANCE DE CARGAS DE TRABAJO_JONATHAN MEDINA.xlsx")
BALANCE_SHEET_MUESTREO = "MUESTREO"
LISTA_MAESTRA_SHEET = "Lista Maestra"


# ----------------------------------------

# ----------------- Funciones de tiempo del algoritmo de análisis -----------------
def convertir_tiempo_a_minutos(tiempo):
    try:
        if pd.isna(tiempo) or tiempo == '':
            return 0.0
        if isinstance(tiempo, pd.Timedelta) or isinstance(tiempo, timedelta):
            return tiempo.total_seconds() / 60.0
        if isinstance(tiempo, (int, float, np.number)):
            return float(tiempo) * 60.0  # interpreta como horas
        if isinstance(tiempo, time):
            return 0.0
        s = str(tiempo).strip()
        if ':' in s:
            partes = [p.split('.')[0] for p in s.split(':')]
            if len(partes) == 3:
                h, m, sec = int(partes[0]), int(partes[1]), int(partes[2])
                return h * 60 + m + sec / 60.0
            elif len(partes) == 2:
                m, sec = int(partes[0]), int(partes[1])
                return m + sec / 60.0
            else:
                try:
                    val = float(partes[0])
                    return val * 60.0 if val <= 24 else val
                except:
                    return 0.0
        try:
            f = float(s)
            return f * 60.0
        except:
            pass
        try:
            td = pd.to_timedelta(s)
            return td.total_seconds() / 60.0
        except:
            pass
        print(f"[WARN] No se pudo convertir tiempo: '{tiempo}' -> 0")
        return 0.0
    except Exception as e:
        print(f"[ERROR] convertir_tiempo_a_minutos('{tiempo}') -> {e}")
        return 0.0


def minutos_a_formato_tiempo(minutos):
    try:
        if pd.isna(minutos):
            minutos = 0.0
        minutos = float(minutos)
        horas = int(minutos // 60)
        mins = int(minutos % 60)
        segs = int(round((minutos - int(minutos)) * 60))
        if segs >= 60:
            mins += segs // 60
            segs = segs % 60
        if mins >= 60:
            horas += mins // 60
            mins = mins % 60
        return f"{horas}:{mins:02d}:{segs:02d}"
    except:
        return "0:00:00"


def calcular_tiempo_desde_intervalo(inicio, fin):
    try:
        if (pd.isna(inicio) or inicio == '') and (pd.isna(fin) or fin == ''):
            return 0.0

        def _to_dt(v):
            if pd.isna(v) or v == '':
                return None
            if isinstance(v, (pd.Timestamp, datetime)):
                return pd.to_datetime(v)
            try:
                dt = pd.to_datetime(str(v), errors='coerce')
                if not pd.isna(dt):
                    return dt
            except:
                return None
            return None

        dt_inicio = _to_dt(inicio)
        dt_fin = _to_dt(fin)
        if dt_inicio is not None and dt_fin is not None:
            if dt_fin < dt_inicio:
                dt_fin = dt_fin + pd.Timedelta(days=1)
            diff = dt_fin - dt_inicio
            return diff.total_seconds() / 60.0
        return 0.0
    except Exception as e:
        print(f"[WARN] calcular_tiempo_desde_intervalo('{inicio}','{fin}') -> {e}")
        return 0.0


def crear_datos_desde_excel(ruta_archivo, sheet_name=BALANCE_SHEET_MUESTREO):
    """Versión adaptada para leer desde la hoja MUESTREO específicamente"""
    if not os.path.exists(ruta_archivo):
        raise FileNotFoundError(f"No existe: {ruta_archivo}")
    try:
        df = pd.read_excel(ruta_archivo, sheet_name=sheet_name, engine='openpyxl', header=2, usecols="B:R")
    except Exception:
        try:
            df = pd.read_excel(ruta_archivo, sheet_name=sheet_name, engine='openpyxl', header=2)
        except Exception as e:
            st.error(f"Error leyendo la hoja {sheet_name}: {e}")
            # Intentar con openpyxl directamente para detectar header
            wb = load_workbook(ruta_archivo, read_only=True, data_only=True)
            ws = wb[sheet_name]
            data = []
            # Buscar el header
            header_row = None
            for r in range(1, min(ws.max_row, 10) + 1):
                row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 20) + 1)]
                row_str = " ".join(str(v) for v in row_vals if v)
                if "ACTIV" in row_str.upper() and "ACCION" in row_str.upper():
                    header_row = r
                    break

            if header_row:
                # Leer datos
                for r in range(header_row + 1, ws.max_row + 1):
                    row_data = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 20) + 1)]
                    if any(v not in (None, "") for v in row_data):
                        data.append(row_data)

                # Crear DataFrame
                df = pd.DataFrame(data)
                # Asignar nombres de columnas basados en header_row
                if header_row <= ws.max_row:
                    header_vals = [ws.cell(row=header_row, column=c).value for c in
                                   range(1, min(ws.max_column, 20) + 1)]
                    df.columns = header_vals[:len(df.columns)]
            else:
                df = pd.DataFrame()
            wb.close()

    # Limpiar nombres de columnas
    df.columns = [str(c).strip().upper() if isinstance(c, str) else str(c) for c in df.columns]

    canonic = {
        'EMPRESA': 'EMPRESA',
        'DEPARTAMENTO': 'DEPARTAMENTO',
        'CARGO': 'CARGO',
        'EVALUADO': 'EVALUADO',
        'FECHA': 'FECHA',
        'DEPARTAMENTO SOPORTE': 'DEPARTAMENTO SOPORTE',
        'CARGO SOPORTE': 'CARGO SOPORTE',
        'ACTIVIDAD': 'ACTIVIDAD',
        'PROCESO': 'PROCESO',
        'ACCIONES': 'ACCIONES',
        'TIPO DE ANALISIS': 'TIPO DE ANALISIS',
        'DEPARTAMENTO ENC': 'DEPARTAMENTO ENC',
        'TIEMPO INICIO': 'TIEMPO INICIO',
        'TIEMPO FIN': 'TIEMPO FIN',
        'TIEMPO': 'TIEMPO',
        'OBSERVACIONES': 'OBSERVACIONES',
        'OPORTUNIDADES DE MEJORA': 'OPORTUNIDADES DE MEJORA'
    }

    for col in canonic.values():
        if col not in df.columns:
            df[col] = pd.NA

    df['FECHA_ORIG'] = df['FECHA']
    try:
        df['FECHA'] = pd.to_datetime(df['FECHA'], errors='coerce')
        mask_nat = df['FECHA'].isna()
        if mask_nat.any():
            df.loc[mask_nat, 'FECHA'] = df.loc[mask_nat, 'FECHA_ORIG'].astype(str)
    except:
        df['FECHA'] = df['FECHA_ORIG'].astype(str)

    minutos = []
    for _, row in df.iterrows():
        tiempo_col = row.get('TIEMPO', pd.NA)
        tiempo_min = 0.0
        if not (pd.isna(tiempo_col) or str(tiempo_col).strip() == ''):
            tiempo_min = convertir_tiempo_a_minutos(tiempo_col)
        else:
            inicio = row.get('TIEMPO INICIO', pd.NA)
            fin = row.get('TIEMPO FIN', pd.NA)
            tiempo_min = calcular_tiempo_desde_intervalo(inicio, fin)
            if tiempo_min == 0:
                if not (pd.isna(inicio) or str(inicio).strip() == ''):
                    posible = convertir_tiempo_a_minutos(inicio)
                    if posible > 0:
                        tiempo_min = posible
                elif not (pd.isna(fin) or str(fin).strip() == ''):
                    posible = convertir_tiempo_a_minutos(fin)
                    if posible > 0:
                        tiempo_min = posible
        minutos.append(float(tiempo_min))
    df['TIEMPO_MINUTOS'] = minutos

    for col in ['EMPRESA', 'DEPARTAMENTO', 'CARGO', 'EVALUADO', 'PROCESO', 'ACTIVIDAD', 'ACCIONES',
                'OPORTUNIDADES DE MEJORA', 'OBSERVACIONES']:
        if col in df.columns:
            df[col] = df[col].fillna('Sin especificar').astype(str).str.strip()

    df['PROCESO_ACTIVIDAD'] = df['PROCESO'] + ' - ' + df['ACTIVIDAD']
    df['PROCESO_ACCIONES'] = df['PROCESO'] + ' - ' + df['ACCIONES']

    if 'SUBPROCESO' in df.columns and df['SUBPROCESO'].notna().any():
        df['PROCESO_SUBPROCESO'] = df['PROCESO'] + ' - ' + df['SUBPROCESO'].fillna('')
    else:
        df['PROCESO_SUBPROCESO'] = df['PROCESO_ACTIVIDAD']

    df['PROCESO_ACT_ACC'] = df['PROCESO'] + ' | ' + df['ACTIVIDAD'] + ' | ' + df['ACCIONES']
    df['EMP_DEP_CARGO'] = df['EMPRESA'] + ' | ' + df['DEPARTAMENTO'] + ' | ' + df['CARGO']

    df_val = df[df['TIEMPO_MINUTOS'] > 0].copy()
    return df_val


# ----------------- Gráficos interactivos con Plotly -----------------
def crear_grafico_barras_fecha_proceso_interactivo(df):
    """Gráfico interactivo de distribución por fecha y proceso"""
    if df.empty:
        return None

    # Preparar datos
    if df['FECHA'].dtype == 'datetime64[ns]' or isinstance(df['FECHA'].iloc[0], pd.Timestamp):
        df['FECHA_STR'] = df['FECHA'].dt.strftime('%Y-%m-%d')
    else:
        df['FECHA_STR'] = df['FECHA'].astype(str)

    resumen = df.groupby(['FECHA_STR', 'PROCESO'])['TIEMPO_MINUTOS'].sum().reset_index()
    resumen = resumen.sort_values(['FECHA_STR', 'TIEMPO_MINUTOS'], ascending=[True, False])

    # Calcular porcentajes
    tiempo_total = df['TIEMPO_MINUTOS'].sum()
    resumen['PORCENTAJE'] = (resumen['TIEMPO_MINUTOS'] / tiempo_total) * 100
    resumen['TIEMPO_FORMATO'] = resumen['TIEMPO_MINUTOS'].apply(minutos_a_formato_tiempo)

    # Crear etiquetas para hover
    resumen['TEXTO_HOVER'] = resumen.apply(
        lambda row: f"<b>Fecha:</b> {row['FECHA_STR']}<br>"
                    f"<b>Proceso:</b> {row['PROCESO']}<br>"
                    f"<b>Tiempo:</b> {row['TIEMPO_FORMATO']}<br>"
                    f"<b>Porcentaje:</b> {row['PORCENTAJE']:.1f}%<br>"
                    f"<b>Minutos:</b> {row['TIEMPO_MINUTOS']:.1f}",
        axis=1
    )

    # Crear gráfico con Plotly
    fig = px.bar(
        resumen,
        x='TIEMPO_MINUTOS',
        y='FECHA_STR',
        color='PROCESO',
        orientation='h',
        hover_name='PROCESO',
        hover_data={'TEXTO_HOVER': True, 'TIEMPO_MINUTOS': False, 'FECHA_STR': False, 'PROCESO': False},
        labels={'TIEMPO_MINUTOS': 'Tiempo (minutos)', 'FECHA_STR': 'Fecha'},
        title='Distribución de Tiempo por Fecha y Proceso',
        height=600
    )

    # Mejorar layout
    fig.update_layout(
        showlegend=True,
        legend_title_text='Procesos',
        hovermode='closest',
        plot_bgcolor='white',
        xaxis=dict(
            title='Tiempo (minutos)',
            gridcolor='lightgray'
        ),
        yaxis=dict(
            title='Fecha',
            categoryorder='category ascending',
            gridcolor='lightgray'
        ),
        margin=dict(l=50, r=50, t=80, b=50)
    )

    # Mejorar tooltips
    fig.update_traces(
        hovertemplate='%{customdata[0]}<extra></extra>'
    )

    return fig


def crear_grafico_proceso_pie_interactivo(df, fecha=None):
    """Gráfico de pastel interactivo para procesos"""
    if df.empty:
        return None

    # Filtrar por fecha si se especifica
    if fecha:
        if df['FECHA'].dtype == 'datetime64[ns]' or isinstance(df['FECHA'].iloc[0], pd.Timestamp):
            fecha_str = pd.to_datetime(fecha).strftime('%Y-%m-%d')
            mask = df['FECHA'].dt.strftime('%Y-%m-%d') == fecha_str
        else:
            mask = df['FECHA'].astype(str) == str(fecha)
        df_filtrado = df[mask].copy()
        if df_filtrado.empty:
            return None
    else:
        df_filtrado = df.copy()

    # Agrupar por proceso
    resumen = df_filtrado.groupby('PROCESO')['TIEMPO_MINUTOS'].sum().reset_index()
    tiempo_total = df_filtrado['TIEMPO_MINUTOS'].sum()
    resumen['PORCENTAJE'] = (resumen['TIEMPO_MINUTOS'] / tiempo_total) * 100

    # Agrupar categorías pequeñas en "OTROS"
    threshold = 3.0  # Porcentaje mínimo
    mayores = resumen[resumen['PORCENTAJE'] >= threshold].copy()
    menores = resumen[resumen['PORCENTAJE'] < threshold].copy()

    if len(menores) > 0:
        otros_sum = menores['TIEMPO_MINUTOS'].sum()
        otros_pct = menores['PORCENTAJE'].sum()
        agrupado = pd.concat([
            mayores,
            pd.DataFrame([{'PROCESO': 'OTROS', 'TIEMPO_MINUTOS': otros_sum, 'PORCENTAJE': otros_pct}])
        ], ignore_index=True)
    else:
        agrupado = mayores.copy()

    # Ordenar
    agrupado = agrupado.sort_values('TIEMPO_MINUTOS', ascending=False)

    # Crear texto para hover
    agrupado['TIEMPO_FORMATO'] = agrupado['TIEMPO_MINUTOS'].apply(minutos_a_formato_tiempo)
    agrupado['TEXTO_HOVER'] = agrupado.apply(
        lambda row: f"<b>Proceso:</b> {row['PROCESO']}<br>"
                    f"<b>Tiempo:</b> {row['TIEMPO_FORMATO']}<br>"
                    f"<b>Porcentaje:</b> {row['PORCENTAJE']:.1f}%<br>"
                    f"<b>Minutos:</b> {row['TIEMPO_MINUTOS']:.1f}",
        axis=1
    )

    # Crear gráfico de pastel
    fig = px.pie(
        agrupado,
        values='TIEMPO_MINUTOS',
        names='PROCESO',
        hover_name='PROCESO',
        hover_data={'TEXTO_HOVER': True, 'TIEMPO_MINUTOS': False, 'PORCENTAJE': False},
        title=f'Distribución de Tiempo por Proceso{f" - {fecha}" if fecha else ""}',
        height=500
    )

    # Mejorar layout
    fig.update_layout(
        showlegend=True,
        legend_title_text='Procesos',
        hovermode='closest',
        margin=dict(l=20, r=20, t=60, b=20)
    )

    fig.update_traces(
        textposition='inside',
        textinfo='percent+label',
        hovertemplate='%{customdata[0]}<extra></extra>',
        textfont_size=12
    )

    return fig


def crear_grafico_subprocesos_interactivo(df, fecha=None, top_n=10):
    """Gráfico de barras interactivo para subprocesos"""
    if df.empty:
        return None

    # Filtrar por fecha si se especifica
    if fecha:
        if df['FECHA'].dtype == 'datetime64[ns]' or isinstance(df['FECHA'].iloc[0], pd.Timestamp):
            fecha_str = pd.to_datetime(fecha).strftime('%Y-%m-%d')
            mask = df['FECHA'].dt.strftime('%Y-%m-%d') == fecha_str
        else:
            mask = df['FECHA'].astype(str) == str(fecha)
        df_filtrado = df[mask].copy()
        if df_filtrado.empty:
            return None
    else:
        df_filtrado = df.copy()

    # Usar PROCESO_SUBPROCESO
    resumen = df_filtrado.groupby('PROCESO_SUBPROCESO')['TIEMPO_MINUTOS'].sum().reset_index()
    tiempo_total = df_filtrado['TIEMPO_MINUTOS'].sum()

    # Ordenar y tomar top N
    resumen = resumen.sort_values('TIEMPO_MINUTOS', ascending=False).head(top_n)
    resumen = resumen.sort_values('TIEMPO_MINUTOS', ascending=True)  # Para barras horizontales

    # Calcular porcentajes y formato
    resumen['PORCENTAJE'] = (resumen['TIEMPO_MINUTOS'] / tiempo_total) * 100
    resumen['TIEMPO_FORMATO'] = resumen['TIEMPO_MINUTOS'].apply(minutos_a_formato_tiempo)

    # Crear texto para hover
    resumen['TEXTO_HOVER'] = resumen.apply(
        lambda row: f"<b>Subproceso/Actividad:</b> {row['PROCESO_SUBPROCESO']}<br>"
                    f"<b>Tiempo:</b> {row['TIEMPO_FORMATO']}<br>"
                    f"<b>Porcentaje:</b> {row['PORCENTAJE']:.1f}%<br>"
                    f"<b>Minutos:</b> {row['TIEMPO_MINUTOS']:.1f}",
        axis=1
    )

    # Crear gráfico de barras
    fig = px.bar(
        resumen,
        x='TIEMPO_MINUTOS',
        y='PROCESO_SUBPROCESO',
        orientation='h',
        hover_name='PROCESO_SUBPROCESO',
        hover_data={'TEXTO_HOVER': True, 'TIEMPO_MINUTOS': False, 'PORCENTAJE': False},
        labels={'TIEMPO_MINUTOS': 'Tiempo (minutos)', 'PROCESO_SUBPROCESO': 'Subproceso/Actividad'},
        title=f'Top {top_n} Subprocesos por Tiempo{f" - {fecha}" if fecha else ""}',
        height=500
    )

    # Mejorar layout
    fig.update_layout(
        showlegend=False,
        hovermode='closest',
        plot_bgcolor='white',
        xaxis=dict(
            title='Tiempo (minutos)',
            gridcolor='lightgray'
        ),
        yaxis=dict(
            title='',
            categoryorder='total ascending',
            tickmode='linear'
        ),
        margin=dict(l=150, r=50, t=80, b=50)
    )

    # Añadir etiquetas de valor
    for i, row in resumen.iterrows():
        fig.add_annotation(
            x=row['TIEMPO_MINUTOS'],
            y=row['PROCESO_SUBPROCESO'],
            text=row['TIEMPO_FORMATO'],
            showarrow=False,
            xanchor='left',
            xshift=10,
            font=dict(size=10)
        )

    fig.update_traces(
        hovertemplate='%{customdata[0]}<extra></extra>'
    )

    return fig


def crear_grafico_proceso_acciones_interactivo(df):
    """Gráfico interactivo de proceso vs acciones"""
    if df.empty:
        return None

    resumen = df.groupby(['PROCESO', 'ACCIONES'])['TIEMPO_MINUTOS'].sum().reset_index()
    tiempo_total = df['TIEMPO_MINUTOS'].sum()
    resumen['PORCENTAJE'] = (resumen['TIEMPO_MINUTOS'] / tiempo_total) * 100
    resumen['TIEMPO_FORMATO'] = resumen['TIEMPO_MINUTOS'].apply(minutos_a_formato_tiempo)

    # Crear etiqueta combinada
    resumen['PROCESO_ACCION'] = resumen['PROCESO'] + ' - ' + resumen['ACCIONES'].str.slice(0, 30)

    # Ordenar
    resumen = resumen.sort_values('TIEMPO_MINUTOS', ascending=True)

    # Crear texto para hover
    resumen['TEXTO_HOVER'] = resumen.apply(
        lambda row: f"<b>Proceso:</b> {row['PROCESO']}<br>"
                    f"<b>Acciones:</b> {row['ACCIONES']}<br>"
                    f"<b>Tiempo:</b> {row['TIEMPO_FORMATO']}<br>"
                    f"<b>Porcentaje:</b> {row['PORCENTAJE']:.1f}%<br>"
                    f"<b>Minutos:</b> {row['TIEMPO_MINUTOS']:.1f}",
        axis=1
    )

    # Crear gráfico de barras
    fig = px.bar(
        resumen,
        x='TIEMPO_MINUTOS',
        y='PROCESO_ACCION',
        orientation='h',
        hover_name='PROCESO_ACCION',
        hover_data={'TEXTO_HOVER': True, 'TIEMPO_MINUTOS': False, 'PORCENTAJE': False},
        color='PROCESO',
        labels={'TIEMPO_MINUTOS': 'Tiempo (minutos)', 'PROCESO_ACCION': 'Proceso - Acciones'},
        title='Distribución de Tiempo por Proceso-Acciones',
        height=max(400, len(resumen) * 25)
    )

    # Mejorar layout
    fig.update_layout(
        showlegend=True,
        legend_title_text='Procesos',
        hovermode='closest',
        plot_bgcolor='white',
        xaxis=dict(
            title='Tiempo (minutos)',
            gridcolor='lightgray'
        ),
        yaxis=dict(
            title='',
            categoryorder='total ascending',
            tickmode='linear'
        ),
        margin=dict(l=200, r=50, t=80, b=50)
    )

    fig.update_traces(
        hovertemplate='%{customdata[0]}<extra></extra>'
    )

    return fig


def generar_resumen_estadistico_streamlit(df):
    """Versión adaptada de resumen estadístico para Streamlit"""
    if df.empty:
        st.warning("No hay datos para generar resumen estadístico.")
        return

    tiempo_total_min = df['TIEMPO_MINUTOS'].sum()
    tiempo_total_formato = minutos_a_formato_tiempo(tiempo_total_min)
    tiempo_promedio_min = df['TIEMPO_MINUTOS'].mean()
    tiempo_promedio_formato = minutos_a_formato_tiempo(tiempo_promedio_min)

    # Crear métricas
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total de Registros", len(df))
    with col2:
        st.metric("Tiempo Total", tiempo_total_formato)
    with col3:
        st.metric("Tiempo Promedio", tiempo_promedio_formato)
    with col4:
        st.metric("Registros por Día", f"{len(df) / max(len(df['FECHA'].unique()), 1):.1f}")

    # Mostrar distribución por fecha
    st.subheader("📅 Distribución por Fecha")
    fecha_col = 'FECHA_STR' if 'FECHA_STR' in df.columns else 'FECHA'
    resumen_fecha = df.groupby(fecha_col)['TIEMPO_MINUTOS'].agg(['count', 'sum', 'mean']).reset_index()

    for _, row in resumen_fecha.iterrows():
        tiempo_total_fecha = minutos_a_formato_tiempo(row['sum'])
        tiempo_promedio_fecha = minutos_a_formato_tiempo(row['mean'])
        porcentaje = (row['sum'] / tiempo_total_min) * 100 if tiempo_total_min > 0 else 0

        st.write(
            f"**{row[fecha_col]}**: {row['count']} actividades, {tiempo_total_fecha} total ({porcentaje:.1f}%), {tiempo_promedio_fecha} promedio")

    # Mostrar top procesos
    st.subheader("🔄 Top Procesos")
    resumen_proceso = df.groupby('PROCESO')['TIEMPO_MINUTOS'].agg(['count', 'sum']).reset_index()
    resumen_proceso = resumen_proceso.sort_values('sum', ascending=False)

    for _, row in resumen_proceso.head(10).iterrows():
        tiempo_proceso = minutos_a_formato_tiempo(row['sum'])
        porcentaje = (row['sum'] / tiempo_total_min) * 100 if tiempo_total_min > 0 else 0
        st.write(f"**{row['PROCESO']}**: {row['count']} actividades, {tiempo_proceso} ({porcentaje:.1f}%)")


# ---------------- UTIL (Funciones existentes) ----------------
@st.cache_data
def scan_excels(root):
    files = []
    for dirpath, _, filenames in os.walk(root):
        for f in filenames:
            if f.lower().endswith((".xlsx", ".xlsm", ".xls")):
                files.append(os.path.join(dirpath, f))
    files.sort()
    return files


# --- AGREGAR ESTA NUEVA FUNCIÓN AQUÍ ---
@st.cache_data
def cargar_excel_base():
    """Carga el archivo Excel base para añadir nuevos registros."""
    try:
        # NOTA: Añade dtype para evitar el ArrowTypeError local
        df = pd.read_excel(
            BALANCE_PATH,
            sheet_name=BALANCE_SHEET_MUESTREO,
            dtype={'TIEMPO INICIO': str, 'TIEMPO FIN': str}  # 🔧 Solución al error local
        )
        return df
    except Exception as e:
        st.error(f"Error cargando el Excel base: {e}")
        return pd.DataFrame()  # Retorna un DataFrame vacío en caso de error

# Cargar el DataFrame base UNA VEZ al inicio (fuera de cualquier función)
df_base = cargar_excel_base()
# ---------------------------------------------------------

def guardar_y_descargar_registro(vals, df_existente):
    """
    vals: diccionario con los datos del nuevo registro (EMPRESA, ACTIVIDAD, etc.)
    df_existente: DataFrame de pandas cargado desde el Excel original
    """
    # 1. Convertir el nuevo registro en un DataFrame
    nuevo_df = pd.DataFrame([vals])

    # 2. Concatenarlo con los datos existentes
    df_actualizado = pd.concat([df_existente, nuevo_df], ignore_index=True)

    # 3. Crear un archivo Excel EN MEMORIA
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_actualizado.to_excel(writer, index=False, sheet_name='MUESTREO')
    output.seek(0)  # Ir al inicio del archivo en memoria

    # 4. Ofrecer el botón de descarga en Streamlit
    st.download_button(
        label="📥 DESCARGAR EXCEL ACTUALIZADO",
        data=output,
        file_name=f"BALANCE_ACTUALIZADO_{datetime.now().date()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return True


def find_header_row_in_sheet(df):
    """Heurística para detectar fila de encabezado en un DataFrame sin header."""
    for i, row in df.iterrows():
        joined = " ".join([str(x).upper() for x in row.tolist() if pd.notna(x)])
        if (("FUNCIÓN" in joined or "FUNCION" in joined or "ACTIVIDAD" in joined) and
                ("ACCIONES" in joined or "ACCIÓN" in joined) and
                ("PROCESO" in joined or "PROCESOS" in joined)):
            return i
    return None


def find_header_in_worksheet(ws):
    """
    Busca fila de encabezado en worksheet y devuelve (header_row, mapping).
    Esta versión prioriza columnas 'DEPARTAMENTO' y 'CARGO' limpias frente a
    columnas con sufijos como 'SOPORTE' o 'ENC'.
    """

    def norm(s):
        return "" if s is None else " ".join(str(s).upper().strip().split())

    for r in range(1, ws.max_row + 1):
        row_vals = [norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        # condición heurística: fila con ACTIVIDAD/FUNCIÓN y ACCIONES
        if any(("ACTIV" in v or "FUNCION" in v or "FUNCIÓN" in v) for v in row_vals) and any(
                ("ACCION" in v or "ACCIÓN" in v) for v in row_vals):
            mapping = {}
            # guardaremos también variantes (enc, soporte) pero preferiremos la versión "limpia"
            for idx, v in enumerate(row_vals, start=1):
                if not v:
                    continue

                # EMPRESA
                if "EMPRES" in v and "EMPRESA" not in mapping:
                    mapping["EMPRESA"] = idx

                # ACTIVIDAD
                if ("ACTIV" in v or "FUNCI" in v) and "ACTIVIDAD" not in mapping:
                    mapping["ACTIVIDAD"] = idx

                # ACCIONES
                if ("ACCION" in v or "ACCIÓN" in v) and "ACCIONES" not in mapping:
                    mapping["ACCIONES"] = idx

                # PROCESO
                if "PROCES" in v and "PROCESO" not in mapping:
                    mapping["PROCESO"] = idx

                # FECHA
                if v == "FECHA" or "FECHA" in v:
                    if "FECHA" not in mapping:
                        mapping["FECHA"] = idx

                # TIEMPOS
                if "TIEMPO INICIO" in v or "TIEMPO_INICIO" in v or v == "INICIO":
                    if "TIEMPO INICIO" not in mapping:
                        mapping["TIEMPO INICIO"] = idx
                if "TIEMPO FIN" in v or "TIEMPO_FIN" in v or v == "FIN":
                    if "TIEMPO FIN" not in mapping:
                        mapping["TIEMPO FIN"] = idx
                if v == "TIEMPO" or v.endswith("TIEMPO"):
                    if "TIEMPO" not in mapping:
                        mapping["TIEMPO"] = idx

                # EVALUADO
                if "EVALUAD" in v or "EVALUADOR" in v or "EVALUADO" in v:
                    if "EVALUADO" not in mapping:
                        mapping["EVALUADO"] = idx

                # DEPARTAMENTO: distinguir variantes
                if "DEPARTAM" in v or "DEPARTAMENTO" in v or "DEPART" in v:
                    # Si el encabezado contiene ENC / ENC. / ENCARG -> guardamos como variante
                    if any(x in v for x in ("ENC", "ENC.", "ENCARG", "ENCARG.")):
                        if "DEPARTAMENTO ENC" not in mapping:
                            mapping["DEPARTAMENTO ENC"] = idx
                    # Si contiene SOPORTE -> variante 'DEPARTAMENTO SOPORTE'
                    elif "SOPORTE" in v:
                        if "DEPARTAMENTO SOPORTE" not in mapping:
                            mapping["DEPARTAMENTO SOPORTE"] = idx
                    else:
                        # preferimos la columna limpia
                        if "DEPARTAMENTO" not in mapping:
                            mapping["DEPARTAMENTO"] = idx

                # CARGO: distinguir variantes (evitar 'CARGO SOPORTE')
                if "CARGO" in v:
                    if "SOPORTE" in v:
                        if "CARGO SOPORTE" not in mapping:
                            mapping["CARGO SOPORTE"] = idx
                    elif any(x in v for x in ("ENC", "ENC.", "ENCARG")):
                        if "CARGO ENC" not in mapping:
                            mapping["CARGO ENC"] = idx
                    else:
                        if "CARGO" not in mapping:
                            mapping["CARGO"] = idx

                # TIPO DE ANALISIS
                if "TIPO" in v and ("ANALISIS" in v or "ANÁLISIS" in v or "ANALY" in v or "TYPE" in v):
                    if "TIPO DE ANALISIS" not in mapping:
                        mapping["TIPO DE ANALISIS"] = idx

            # Después del escaneo: si existe variante limpia, preferirla. (ya se insertó en orden preferencial)
            # Asegurar que al menos ACTIVIDAD y ACCIONES existan
            if "ACTIVIDAD" in mapping and "ACCIONES" in mapping:
                return r, mapping
    return None, None


def build_lista_maestra_from_table(df_table, empresa_from_filename="TSER"):
    """Construye DF con columnas EMPRESA/ACTIVIDAD/ACCIONES/PROCESO.
       Rellena EMPRESA vacía con empresa_from_filename por defecto.
    """
    if df_table is None or df_table.shape[0] == 0:
        return pd.DataFrame(columns=["EMPRESA", "ACTIVIDAD", "ACCIONES", "PROCESO"])
    upper_map = {c.upper(): c for c in df_table.columns}
    actividad_col = None
    for cand in ("FUNCIÓN", "FUNCION", "ACTIVIDAD", "ACTIVIDADES"):
        for u, c in upper_map.items():
            if cand in u:
                actividad_col = c;
                break
        if actividad_col: break
    acciones_col = None
    for cand in ("ACCIONES", "ACCIÓN", "ACCION"):
        for u, c in upper_map.items():
            if cand in u:
                acciones_col = c;
                break
        if acciones_col: break
    proceso_col = None
    for cand in ("PROCESO", "PROCESOS"):
        for u, c in upper_map.items():
            if cand in u:
                proceso_col = c;
                break
        if proceso_col: break
    empresa_col = None
    for cand in ("EMPRESA", "COMPANY"):
        for u, c in upper_map.items():
            if cand in u:
                empresa_col = c;
                break
        if empresa_col: break

    res = pd.DataFrame()
    if empresa_col:
        res["EMPRESA"] = df_table[empresa_col].fillna("").astype(str)
    else:
        res["EMPRESA"] = empresa_from_filename
    res["ACTIVIDAD"] = df_table[actividad_col] if actividad_col and actividad_col in df_table.columns else ""
    res["ACCIONES"] = df_table[acciones_col] if acciones_col and acciones_col in df_table.columns else ""
    res["PROCESO"] = df_table[proceso_col] if proceso_col and proceso_col in df_table.columns else ""
    res = res.fillna("").astype(str)
    # Rellenar vacíos con default
    res["EMPRESA"] = res["EMPRESA"].apply(lambda x: empresa_from_filename if str(x).strip() == "" else x)
    res = res[res["ACTIVIDAD"].str.strip() != ""]
    res = res.drop_duplicates(subset=["ACTIVIDAD", "ACCIONES", "PROCESO"])
    res = res.reset_index(drop=True)
    return res


def write_edited_table_back_to_source(path, sheet_name, header_row_idx_zero_based, df_edited):
    """
    Escribe la tabla editada (headers + data) de vuelta al libro origen.
    header_row_idx_zero_based = índice de fila (0-based) detectado por pandas cuando leímos la hoja.
    Este valor se obtiene al detectar encabezado en la lectura inicial.
    """
    try:
        wb = load_workbook(path)
    except Exception as e:
        return False, f"No se pudo abrir archivo fuente: {e}"
    if sheet_name not in wb.sheetnames:
        wb.close();
        return False, "No existe la hoja indicada en el archivo fuente."
    ws = wb[sheet_name]
    # Excel row (1-based) del header original
    header_row_excel = header_row_idx_zero_based + 1
    # escribir encabezado (df_edited.columns)
    for j, colname in enumerate(list(df_edited.columns), start=1):
        ws.cell(row=header_row_excel, column=j).value = colname
    # escribir filas a partir de header_row_excel + 1
    start_row = header_row_excel + 1
    # borrar bloque antiguo: opcionalmente limpiar (es peligroso borrar mucho), aquí sobreescribimos filas existentes y limpiamos filas extra hasta len(df)
    # Escribimos cada fila:
    for i, row in df_edited.iterrows():
        for j, colname in enumerate(list(df_edited.columns), start=1):
            ws.cell(row=start_row + i, column=j).value = row[colname]
    wb.save(path)
    wb.close()
    return True, "Hoja fuente actualizada."


def load_lista_maestra_from_balance(path, sheet_name):
    cols = ["EMPRESA", "ACTIVIDAD", "ACCIONES", "PROCESO"]
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return pd.DataFrame(columns=cols)
    if sheet_name not in wb.sheetnames:
        wb.close();
        return pd.DataFrame(columns=cols)
    ws = wb[sheet_name]
    header_row, mapping = find_header_in_worksheet(ws)
    data = []
    if header_row and mapping:
        r = header_row + 1
        while r <= ws.max_row:
            val_act = ws.cell(row=r, column=mapping["ACTIVIDAD"]).value if mapping.get("ACTIVIDAD") else None
            if val_act is None or str(val_act).strip() == "":
                break
            row_vals = []
            for k in cols:
                col_idx = mapping.get(k)
                v = "" if col_idx is None or ws.cell(row=r, column=col_idx).value is None else str(
                    ws.cell(row=r, column=col_idx).value)
                row_vals.append(v)
            data.append(row_vals)
            r += 1
    wb.close()
    if not data:
        return pd.DataFrame(columns=cols)
    df = pd.DataFrame(data, columns=cols)
    df = df.drop_duplicates(subset=["ACTIVIDAD"])
    df = df.reset_index(drop=True)
    return df


def append_or_update_lista_maestra(path, sheet_name, df_new):
    """
    Agrega nuevas actividades y actualiza existentes.
    Versión con manejo de errores para archivos compartidos.
    """
    if df_new is None or df_new.empty:
        return False, "No hay datos nuevos."

    max_intentos = 3
    intento = 0

    while intento < max_intentos:
        try:
            wb = load_workbook(path)
            break
        except Exception as e:
            intento += 1
            if intento == max_intentos:
                return False, f"No se pudo abrir BALANCE después de {max_intentos} intentos: {e}"
            import time
            time.sleep(1)  # Esperar 1 segundo antes de reintentar

    try:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        header_row, mapping = find_header_in_worksheet(ws)
        added = 0
        updated = 0

        if header_row and mapping:
            # Leer actividades existentes
            existing_map = {}
            r = header_row + 1
            while r <= ws.max_row:
                val_act = ws.cell(row=r, column=mapping["ACTIVIDAD"]).value if mapping.get("ACTIVIDAD") else None
                if val_act is None or str(val_act).strip() == "":
                    break
                existing_map[str(val_act).strip()] = r
                r += 1

            insert_row = r

            # Procesar cada fila nueva
            for _, row in df_new.iterrows():
                actividad = str(row.get("ACTIVIDAD", "")).strip()
                if actividad == "":
                    continue

                if actividad in existing_map:
                    # ACTUALIZAR SOLO SI HAY CAMBIOS REALES
                    rtarget = existing_map[actividad]
                    changed = False

                    for colname in ("EMPRESA", "ACCIONES", "PROCESO"):
                        newval = row.get(colname, "")
                        if newval is None:
                            continue
                        newval = str(newval).strip()
                        if newval == "":
                            continue

                        col_idx = mapping.get(colname)
                        if col_idx:
                            current = ws.cell(row=rtarget, column=col_idx).value
                            current_str = "" if current is None else str(current).strip()
                            if current_str != newval:
                                ws.cell(row=rtarget, column=col_idx).value = newval
                                changed = True

                    if changed:
                        updated += 1
                else:
                    # INSERTAR NUEVA
                    for colname in ("EMPRESA", "ACTIVIDAD", "ACCIONES", "PROCESO"):
                        col_idx = mapping.get(colname)
                        val = row.get(colname, "")
                        if col_idx and str(val).strip() != "":
                            ws.cell(row=insert_row, column=col_idx).value = val

                    insert_row += 1
                    added += 1

            wb.save(path)
            wb.close()
            return True, f"{added} filas agregadas, {updated} filas actualizadas."

        else:
            # Crear encabezado en D6 si no existe
            start_col = 4  # D
            start_row = 6
            headers = ["EMPRESA", "ACTIVIDAD", "ACCIONES", "PROCESO"]

            for j, h in enumerate(headers):
                ws.cell(row=start_row, column=start_col + j).value = h

            r = start_row + 1
            added = 0
            for _, row in df_new.iterrows():
                actividad = str(row.get("ACTIVIDAD", "")).strip()
                if actividad == "":
                    continue

                for j, h in enumerate(headers):
                    ws.cell(row=r, column=start_col + j).value = row.get(h, "")

                r += 1
                added += 1

            wb.save(path)
            wb.close()
            return True, f"Encabezado creado y {added} filas agregadas."

    except Exception as e:
        try:
            wb.close()
        except:
            pass
        return False, f"Error al procesar el archivo: {e}"


def set_activity_validation_on_muestreo(path, muestreo_sheet, lista_sheet, lista_col_letter='D', start_row=6,
                                        end_row=190):
    try:
        wb = load_workbook(path)
    except Exception:
        return False, "No se pudo abrir archivo."
    if lista_sheet not in wb.sheetnames or muestreo_sheet not in wb.sheetnames:
        wb.close();
        return False, "Hojas no encontradas."
    ws_m = wb[muestreo_sheet]
    range_ref = f"'{lista_sheet}'!${lista_col_letter}${start_row}:${lista_col_letter}${end_row}"
    dv = DataValidation(type="list", formula1=range_ref, allow_blank=True)
    header_row_idx = None
    act_col = None
    for r in range(1, ws_m.max_row + 1):
        for c in range(1, ws_m.max_column + 1):
            v = ws_m.cell(row=r, column=c).value
            if v and isinstance(v, str) and "ACTIV" in v.upper():
                header_row_idx = r;
                act_col = c;
                break
        if header_row_idx:
            break
    if act_col:
        addr = f"{get_column_letter(act_col)}{header_row_idx + 1}:{get_column_letter(act_col)}{ws_m.max_row}"
        dv.ranges.append(addr)
    else:
        addr = f"H4:H{ws_m.max_row}"
        dv.ranges.append(addr)
    ws_m.add_data_validation(dv)
    wb.save(path);
    wb.close()
    return True, "Validación aplicada."


def load_unique_values_from_muestreo(path, muestreo_sheet, header_names):
    """
    Versión optimizada para extraer históricos de MUESTREO.
    """
    res = {h: [] for h in header_names}
    patrones = {
        "EMPRESA": ["EMPRESA", "COMPANY", "EMPRES"],
        "DEPARTAMENTO": ["DEPARTAMENTO", "DEPARTAM", "DEPT", "DEPARTMENT", "AREA"],
        "CARGO": ["CARGO", "POSITION", "ROLE", "PUESTO"],
        "EVALUADO": ["EVALUADO", "EVALUATE", "PERSON", "NAME", "NOMBRE", "COLABORADOR"],
        "TIPO DE ANALISIS": ["TIPO DE ANALISIS", "TIPO ANALISIS", "TIPO", "ANALISIS", "ANÁLISIS", "TYPE"]
    }

    # 1) INTENTAR CON PANDAS
    try:
        df = pd.read_excel(
            path,
            sheet_name=muestreo_sheet,
            header=2,
            engine='openpyxl',
            dtype=str
        )
        df.columns = [str(col).strip().upper() for col in df.columns]

        for h in header_names:
            col_encontrada = None
            if h in patrones:
                for patron in patrones[h]:
                    for col in df.columns:
                        if patron.upper() in col:
                            col_encontrada = col
                            break
                    if col_encontrada:
                        break

            if not col_encontrada and h in ["EMPRESA", "DEPARTAMENTO", "CARGO", "EVALUADO"]:
                posiciones = {
                    "EMPRESA": 0,
                    "DEPARTAMENTO": 1,
                    "CARGO": 2,
                    "EVALUADO": 3
                }
                if h in posiciones and posiciones[h] < len(df.columns):
                    col_encontrada = df.columns[posiciones[h]]

            if col_encontrada:
                valores = df[col_encontrada].dropna().astype(str).str.strip()
                valores = valores[valores != ""].drop_duplicates().tolist()
                res[h] = sorted(valores)

        return res

    except Exception as e_pandas:
        st.warning(f"Lectura con pandas falló: {str(e_pandas)[:100]}... Usando método alternativo.")

    # 2) FALLBACK CON OPENPYXL
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if muestreo_sheet not in wb.sheetnames:
            wb.close()
            return res

        ws = wb[muestreo_sheet]
        header_row = None
        mapping = {}

        for r in range(1, 11):
            fila_encabezados = []
            for c in range(1, min(ws.max_column, 10) + 1):
                v = ws.cell(row=r, column=c).value
                fila_encabezados.append("" if v is None else str(v).upper().strip())

            encontrados = 0
            temp_mapping = {}
            for c_idx, valor in enumerate(fila_encabezados, start=1):
                if not valor:
                    continue
                for h in header_names:
                    if h in patrones:
                        for patron in patrones[h]:
                            if patron.upper() in valor:
                                if h not in temp_mapping:
                                    temp_mapping[h] = c_idx
                                    encontrados += 1
                                break

            if encontrados >= 3:
                header_row = r
                mapping = temp_mapping
                break

        if not header_row:
            header_row = 3
            mapping = {
                "EMPRESA": 2,
                "DEPARTAMENTO": 3,
                "CARGO": 4,
                "EVALUADO": 5
            }
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=header_row, column=c).value
                if v and isinstance(v, str) and any(
                        palabra in v.upper() for palabra in ["TIPO", "ANALISIS", "ANÁLISIS"]):
                    mapping["TIPO DE ANALISIS"] = c

        if header_row and mapping:
            for h in header_names:
                col_idx = mapping.get(h)
                if not col_idx:
                    continue

                valores_set = set()
                r = header_row + 1
                filas_leidas = 0

                while r <= ws.max_row and filas_leidas < 500:
                    v = ws.cell(row=r, column=col_idx).value
                    if v is not None:
                        v_str = str(v).strip()
                        if v_str and v_str.lower() not in ["", "nan", "none", "null"]:
                            valores_set.add(v_str)
                    r += 1
                    filas_leidas += 1

                res[h] = sorted(list(valores_set))

        wb.close()
        return res

    except Exception as e_openpyxl:
        st.error(f"Error crítico al leer archivo Excel: {str(e_openpyxl)[:100]}")
        return res


def save_record_to_muestreo(vals, activity, tiempo_inicio, tiempo_fin):
    """Guarda o actualiza registro en MUESTREO con mapeo correcto de columnas."""
    try:
        wb = load_workbook(BALANCE_PATH)
    except Exception as e:
        return False, f"No se pudo abrir BALANCE: {e}"

    if BALANCE_SHEET_MUESTREO not in wb.sheetnames:
        wb.close()
        return False, f"No existe hoja {BALANCE_SHEET_MUESTREO}."

    ws = wb[BALANCE_SHEET_MUESTREO]
    header_row, mapping = find_header_in_worksheet(ws)

    if not mapping or not header_row:
        wb.close()
        return False, "No se pudo encontrar el encabezado en la hoja MUESTREO."

    # Buscar primera fila vacía
    check_col = mapping.get("ACTIVIDAD") or mapping.get("EVALUADO") or list(mapping.values())[0]
    r = header_row + 1
    while r <= ws.max_row:
        v = ws.cell(row=r, column=check_col).value
        if v is None or str(v).strip() == "":
            break
        r += 1
    insert_row = r

    # Mapeo CORRECTO de columnas
    col_mappings = {
        "EMPRESA": mapping.get("EMPRESA"),
        "DEPARTAMENTO": mapping.get("DEPARTAMENTO"),
        "CARGO": mapping.get("CARGO"),
        "EVALUADO": mapping.get("EVALUADO"),
        "FECHA": mapping.get("FECHA"),
        "ACTIVIDAD": mapping.get("ACTIVIDAD"),
        "TIEMPO INICIO": mapping.get("TIEMPO INICIO"),
        "TIEMPO FIN": mapping.get("TIEMPO FIN"),
        "TIPO DE ANALISIS": mapping.get("TIPO DE ANALISIS")
    }

    # Si no se encuentra DEPARTAMENTO, buscar alternativas
    if not col_mappings["DEPARTAMENTO"]:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=header_row, column=c).value
            if v and isinstance(v, str):
                v_upper = v.upper()
                if "DEPARTAMENTO" in v_upper and "SOPORTE" not in v_upper:
                    col_mappings["DEPARTAMENTO"] = c
                    break

    # Si no se encuentra CARGO, buscar alternativas
    if not col_mappings["CARGO"]:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=header_row, column=c).value
            if v and isinstance(v, str):
                v_upper = v.upper()
                if "CARGO" in v_upper and "SOPORTE" not in v_upper:
                    col_mappings["CARGO"] = c
                    break

    # Actualizar si coincide last_insert
    if (st.session_state.get("last_insert_row") and
            st.session_state.get("last_insert_activity") == activity and
            st.session_state.get("last_insert_tiempo_inicio") == tiempo_inicio):

        target = st.session_state.get("last_insert_row")
        updated = False

        for field, col_idx in col_mappings.items():
            if col_idx:
                v = vals.get(field, "")
                if field == "FECHA" and isinstance(v, datetime):
                    ws.cell(row=target, column=col_idx).value = v.strftime("%d/%m/%Y")
                    updated = True
                elif v:
                    ws.cell(row=target, column=col_idx).value = v
                    updated = True

        if updated:
            wb.save(BALANCE_PATH)
            wb.close()
            return True, ("updated", target)

    # Insertar nuevo registro
    for field, col_idx in col_mappings.items():
        if col_idx:
            v = vals.get(field, "")
            if field == "FECHA" and isinstance(v, datetime):
                ws.cell(row=insert_row, column=col_idx).value = v.strftime("%d/%m/%Y")
            elif v:
                ws.cell(row=insert_row, column=col_idx).value = v

    wb.save(BALANCE_PATH)
    wb.close()
    return True, ("inserted", insert_row)


# ---------------- UI PRINCIPAL ----------------
st.title("Sistema de Evaluación - Matriz de Funciones → BALANCE (Análisis Integrado)")

# Crear pestañas
tab1, tab2 = st.tabs(["🏢 Gestión de Matriz de Funciones", "📊 Análisis de Datos"])

with tab1:
    # --- Contenido existente de gestión ---
    col1, col2 = st.columns([2, 1])

    with col1:
        st.subheader("1) Selecciona archivo Matriz de funciones")
        files = scan_excels(ROOT_FOLDER)
        if not files:
            st.warning("No encontré archivos .xlsx en la ruta configurada (ROOT_FOLDER).")
        sel_file = st.selectbox("Excel encontrados", options=files, key="sel_file_analisis")
        if sel_file:
            st.markdown(f"**Archivo seleccionado:** `{sel_file}`")
            xls = pd.ExcelFile(sel_file)
            found_sheets = []
            sheet_tables = {}
            for sh in xls.sheet_names:
                try:
                    df_raw = pd.read_excel(xls, sheet_name=sh, header=None, engine="openpyxl")
                except Exception:
                    continue
                header_row_idx = find_header_row_in_sheet(df_raw)
                if header_row_idx is not None:
                    headers = df_raw.iloc[header_row_idx].fillna("").astype(str).tolist()
                    df_table = df_raw.iloc[header_row_idx + 1:].copy()
                    df_table.columns = headers
                    df_table = df_table.dropna(how="all").reset_index(drop=True)
                    found_sheets.append(sh)
                    sheet_tables[sh] = (df_table, header_row_idx)
            if len(found_sheets) == 0:
                st.warning("No encontré el encabezado con la estructura esperada en las hojas de este archivo.")
            else:
                hoja_sel = st.selectbox("Selecciona hoja/pestaña para extraer", options=found_sheets, index=0,
                                        key="hoja_sel_analisis")
                st.write("Hoja seleccionada:", hoja_sel)
                df_table, hdr_idx = sheet_tables[hoja_sel]
                st.markdown("**Tabla completa (edítala si hace falta)**")
                edited = st.data_editor(df_table, num_rows="dynamic", key=f"editor_{hoja_sel}")

                default_empresa_choice = st.selectbox("Default EMPRESA para filas vacías",
                                                      options=["TSER", "SMART", "Mantener carpeta", "Otro"], index=0,
                                                      key="empresa_choice")
                if default_empresa_choice == "Otro":
                    default_empresa_value = st.text_input("Ingresa el nombre de EMPRESA por defecto", value="TSER",
                                                          key="empresa_custom")
                elif default_empresa_choice == "Mantener carpeta":
                    default_empresa_value = os.path.basename(os.path.dirname(sel_file)) or "TSER"
                else:
                    default_empresa_value = default_empresa_choice

                df_lista_extracted = build_lista_maestra_from_table(edited, empresa_from_filename=default_empresa_value)
                st.subheader("Lista maestra (extraída y editada)")
                st.dataframe(df_lista_extracted, width='stretch')

                write_back_checkbox = st.checkbox(
                    "Escribir los cambios también en el archivo fuente (sobrescribir tabla en la hoja seleccionada)",
                    value=False, key="write_back")
                auto_valid = st.checkbox("Aplicar validación automática en hoja MUESTREO (crear lista desplegable)",
                                         value=False, key="auto_valid")
                if st.button("Importar / Actualizar Lista Maestra al BALANCE (añadir/actualizar filas)",
                             key="import_lista"):
                    ok, msg = append_or_update_lista_maestra(BALANCE_PATH, LISTA_MAESTRA_SHEET, df_lista_extracted)
                    if ok:
                        st.success("Lista Maestra actualizada en el archivo BALANCE. " + msg)
                        if write_back_checkbox:
                            ok_src, msg_src = write_edited_table_back_to_source(sel_file, hoja_sel, hdr_idx, edited)
                            if ok_src:
                                st.info("Hoja origen actualizada en archivo fuente.")
                            else:
                                st.warning("No se pudo escribir en archivo fuente: " + msg_src)
                        if auto_valid:
                            okv, msgv = set_activity_validation_on_muestreo(BALANCE_PATH, BALANCE_SHEET_MUESTREO,
                                                                            LISTA_MAESTRA_SHEET)
                            if okv:
                                st.info("Validación aplicada en MUESTREO.")
                            else:
                                st.warning("No se pudo aplicar validación: " + msgv)
                        st.info(
                            "Importación completada. Si no ves los cambios en la sección 2, recarga el navegador (F5).")
                    else:
                        st.error("No se actualizó: " + msg)

    with col2:
        st.subheader("2) Vista previa Lista Maestra en BALANCE")
        df_balance_lista = load_lista_maestra_from_balance(BALANCE_PATH, LISTA_MAESTRA_SHEET)
        if df_balance_lista.empty:
            st.info("No se encontró Lista Maestra en el BALANCE (buscando encabezados).")
        else:
            st.dataframe(df_balance_lista, width='stretch')

    st.markdown("---")
    st.subheader("3) Interfaz para evaluación en tiempo real (evaluado)")

    # cargar historicos desde MUESTREO
    hist = load_unique_values_from_muestreo(BALANCE_PATH, BALANCE_SHEET_MUESTREO,
                                            ["EMPRESA", "DEPARTAMENTO", "CARGO", "EVALUADO", "TIPO DE ANALISIS"])

    # EVALUADO (histórico)
    evaluado_options = ["--Nuevo--"] + hist.get("EVALUADO", [])
    evaluado_sel = st.selectbox("EVALUADO (histórico)", options=evaluado_options, key="evaluado_sel")
    if evaluado_sel == "--Nuevo--":
        evaluado_input = st.text_input("Si no está en la lista, escribe EVALUADO nuevo", value="", key="evaluado_new")
    else:
        evaluado_input = evaluado_sel

    # EMPRESA (histórico)
    empresa_options = ["--Nuevo--"] + hist.get("EMPRESA", [])
    empresa_sel = st.selectbox("EMPRESA (histórico)", options=empresa_options, index=0, key="empresa_sel")
    if empresa_sel == "--Nuevo--":
        empresa_input = st.text_input("EMPRESA (nuevo si aplica)", value="", key="empresa_new")
    else:
        empresa_input = empresa_sel

    # DEPARTAMENTO (histórico)
    departamento_options = ["--Nuevo--"] + hist.get("DEPARTAMENTO", [])
    departamento_sel = st.selectbox("DEPARTAMENTO (histórico)", options=departamento_options, index=0, key="dept_sel")
    if departamento_sel == "--Nuevo--":
        departamento = st.text_input("DEPARTAMENTO (nuevo si aplica)", value="", key="dept_new")
    else:
        departamento = departamento_sel

    # CARGO (histórico)
    cargo_options = ["--Nuevo--"] + hist.get("CARGO", [])
    cargo_sel = st.selectbox("CARGO (histórico)", options=cargo_options, index=0, key="cargo_sel")
    if cargo_sel == "--Nuevo--":
        cargo = st.text_input("CARGO (nuevo si aplica)", value="", key="cargo_new")
    else:
        cargo = cargo_sel

    # Evaluador default (inmutable)
    evaluador = st.text_input("Evaluador (TU NOMBRE)", value="JONATHAN ALEXANDER MEDINA MONTESDEOCA", disabled=True,
                              key="evaluador")

    # ACTIVIDADES desde Lista Maestra
    df_balance_lista = load_lista_maestra_from_balance(BALANCE_PATH, LISTA_MAESTRA_SHEET)
    df_actividades = df_balance_lista["ACTIVIDAD"].tolist() if not df_balance_lista.empty else []

    # session keys
    if "last_actividad" not in st.session_state:
        st.session_state["last_actividad"] = None
    if "tiempo_inicio" not in st.session_state:
        st.session_state["tiempo_inicio"] = None
    if "tiempo_fin" not in st.session_state:
        st.session_state["tiempo_fin"] = None
    if "last_insert_row" not in st.session_state:
        st.session_state["last_insert_row"] = None
    if "last_insert_activity" not in st.session_state:
        st.session_state["last_insert_activity"] = None
    if "last_insert_tiempo_inicio" not in st.session_state:
        st.session_state["last_insert_tiempo_inicio"] = None

    colA, colB, colC = st.columns([2, 2, 2])
    with colA:
        actividad_sel = st.selectbox("Selecciona ACTIVIDAD", options=[""] + df_actividades, key="actividad_sel")
    with colB:
        def infer_tipo_from_muestreo(activity):
            if not activity:
                return ""
            try:
                wb = load_workbook(BALANCE_PATH, read_only=True, data_only=True)
            except Exception:
                return ""
            if BALANCE_SHEET_MUESTREO not in wb.sheetnames:
                wb.close();
                return ""
            ws = wb[BALANCE_SHEET_MUESTREO]
            header_row, mapping = find_header_in_worksheet(ws)
            tipo_val = ""
            if header_row and mapping and mapping.get("TIPO DE ANALISIS"):
                r = header_row + 1
                last_found = None
                while r <= ws.max_row:
                    v_act = ws.cell(row=r, column=mapping.get("ACTIVIDAD")).value if mapping.get("ACTIVIDAD") else None
                    if v_act and str(v_act).strip() == activity:
                        t = ws.cell(row=r, column=mapping.get("TIPO DE ANALISIS")).value
                        if t and str(t).strip() != "":
                            last_found = str(t).strip()
                    r += 1
                if last_found:
                    tipo_val = last_found
            wb.close()
            if tipo_val:
                return tipo_val
            a = activity.strip().upper()
            if "NECESIDADES BIOLÓGICAS" in a:
                return "NECESIDADES BIOLÓGICAS"
            if "ALIMENTACIÓN" in a:
                return "ALIMENTACIÓN"
            return "CON ACTIVIDAD"


        tipo_analisis_val = infer_tipo_from_muestreo(actividad_sel)
        st.text_input("TIPO DE ANALISIS", value=tipo_analisis_val, disabled=True, key="tipo_analisis")
    with colC:
        btn_start = st.button("Iniciar actividad (capturar TIEMPO INICIO)", key="btn_start")
        btn_stop = st.button("Finalizar actividad (capturar TIEMPO FIN)", key="btn_stop")

    if actividad_sel and st.session_state.get("last_actividad") != actividad_sel:
        st.session_state["tiempo_inicio"] = datetime.now()
        st.session_state["tiempo_fin"] = None
        st.session_state["last_actividad"] = actividad_sel

    if btn_start:
        st.session_state["tiempo_inicio"] = datetime.now()
        st.success(f"TIEMPO INICIO fijado: {st.session_state['tiempo_inicio'].strftime('%Y-%m-%d %H:%M:%S')}")
    if btn_stop:
        st.session_state["tiempo_fin"] = datetime.now()
        st.success(f"TIEMPO FIN fijado: {st.session_state['tiempo_fin'].strftime('%Y-%m-%d %H:%M:%S')}")
        auto_save_on_finish = True
    else:
        auto_save_on_finish = False

    # mostrar ACCIONES/PROCESO solo visual
    acciones_val = ""
    proceso_val = ""
    if actividad_sel:
        match = df_balance_lista[df_balance_lista["ACTIVIDAD"] == actividad_sel]
        if not match.empty:
            acciones_val = match.iloc[0].get("ACCIONES", "")
            proceso_val = match.iloc[0].get("PROCESO", "")
    st.text_area("ACCIONES (solo visual)", value=acciones_val, height=120, disabled=True, key="acciones_view")
    st.text_input("PROCESO (solo visual)", value=proceso_val, disabled=True, key="proceso_view")

    st.write("TIEMPO INICIO:", st.session_state.get("tiempo_inicio"))
    st.write("TIEMPO FIN:", st.session_state.get("tiempo_fin"))
    tiempo_horas = ""
    if st.session_state.get("tiempo_inicio") and st.session_state.get("tiempo_fin"):
        delta = st.session_state["tiempo_fin"] - st.session_state["tiempo_inicio"]
        tiempo_horas = delta.total_seconds() / 3600
        st.write(f"TIEMPO (horas): {tiempo_horas:.4f}")

    # Botón Guardar manual
    if st.button("Guardar registro en BALANCE (append en tabla MUESTREO)", key="btn_save"):
        if not actividad_sel:
            st.error("Selecciona una actividad antes de guardar.")
        else:
            # --- CORRECCIÓN DE ZONA HORARIA PARA TODA LA OPERACIÓN ---
            from datetime import datetime
            import pytz

            utc_ahora = datetime.now(pytz.UTC)
            zona_local = pytz.timezone('America/Guayaquil')
            hora_local = utc_ahora.astimezone(zona_local)
            # --------------------------------------------------------

            # Establecer tiempos con hora LOCAL corregida
            if not st.session_state.get("tiempo_fin"):
                st.session_state["tiempo_fin"] = hora_local
                st.success(f"TIEMPO FIN establecido: {hora_local.strftime('%Y-%m-%d %H:%M:%S')}")

            if not st.session_state.get("tiempo_inicio"):
                st.session_state["tiempo_inicio"] = hora_local
                st.success(f"TIEMPO INICIO establecido: {hora_local.strftime('%Y-%m-%d %H:%M:%S')}")

            # Crear diccionario vals con hora LOCAL
            vals = {
                "EMPRESA": empresa_input,
                "DEPARTAMENTO": departamento,
                "CARGO": cargo,
                "EVALUADO": evaluado_input,
                "FECHA": hora_local.strftime("%d/%m/%Y"),
                "ACTIVIDAD": actividad_sel,
                "TIEMPO INICIO": st.session_state.get("tiempo_inicio").strftime("%H:%M:%S") if st.session_state.get(
                    "tiempo_inicio") else "",
                "TIEMPO FIN": st.session_state.get("tiempo_fin").strftime("%H:%M:%S") if st.session_state.get(
                    "tiempo_fin") else "",
                "TIPO DE ANALISIS": tipo_analisis_val
            }

            # ✅ LLAMADA A LA NUEVA FUNCIÓN (reemplaza save_record_to_muestreo)
            if not df_base.empty:
                st.success(
                    "✅ Registro preparado correctamente. Usa el botón de abajo para descargar el Excel actualizado.")
                guardar_y_descargar_registro(vals, df_base)
                # Opcional: Reiniciar estado para nuevo registro
                st.session_state["last_insert_activity"] = actividad_sel
                st.session_state["last_insert_tiempo_inicio"] = st.session_state.get("tiempo_inicio")
            else:
                st.error(
                    "❌ No se pudo cargar el archivo base. Verifica que 'BALANCE DE CARGAS...xlsx' esté en la carpeta /data.")

    # Actualizar el bloque auto_save_on_finish (justo después del botón manual)
    if 'auto_save_on_finish' not in st.session_state:
        st.session_state['auto_save_on_finish'] = False

    if auto_save_on_finish and actividad_sel:
        # Usar la misma lógica de zona horaria
        utc_ahora = datetime.now(pytz.UTC)
        zona_local = pytz.timezone('America/Guayaquil')
        hora_local = utc_ahora.astimezone(zona_local)

        if not st.session_state.get("tiempo_inicio"):
            st.session_state["tiempo_inicio"] = hora_local
        if not st.session_state.get("tiempo_fin"):
            st.session_state["tiempo_fin"] = hora_local

        vals = {
            "EMPRESA": empresa_input,
            "DEPARTAMENTO": departamento,
            "CARGO": cargo,
            "EVALUADO": evaluado_input,
            "FECHA": hora_local.strftime("%d/%m/%Y"),
            "ACTIVIDAD": actividad_sel,
            "TIEMPO INICIO": st.session_state.get("tiempo_inicio").strftime("%H:%M:%S"),
            "TIEMPO FIN": st.session_state.get("tiempo_fin").strftime("%H:%M:%S"),
            "TIPO DE ANALISIS": tipo_analisis_val
        }

        if not df_base.empty:
            st.success("✅ Registro auto-guardado. Descarga el Excel actualizado:")
            guardar_y_descargar_registro(vals, df_base)
            st.session_state["last_insert_activity"] = actividad_sel
            st.session_state["last_insert_tiempo_inicio"] = st.session_state.get("tiempo_inicio")
        else:
            st.error("❌ No se pudo cargar el archivo base para auto-guardado.")

    st.markdown("---")

    # Generar requirements.txt
    if st.button("Generar requirements.txt (pip freeze)", key="btn_req"):
        try:
            txt = subprocess.check_output(["pip", "freeze"], universal_newlines=True)
        except Exception as e:
            st.error(f"No se pudo ejecutar pip freeze: {e}")
            txt = ""
        if txt:
            req_path = os.path.join(os.getcwd(), "requirements.txt")
            with open(req_path, "w", encoding="utf-8") as f:
                f.write(txt)
            st.success("requirements.txt generado en el directorio del proyecto.")
            with open(req_path, "r") as f:
                st.download_button(
                    label="Descargar requirements.txt",
                    data=f.read(),
                    file_name="requirements.txt",
                    mime="text/plain"
                )

    st.caption("Cierra los archivos Excel en Excel antes de ejecutar operaciones que escriben en ellos.")
    st.info("Para detener la app: Ctrl+C en la terminal o Stop en PyCharm Run/Debug.")

with tab2:
    # --- Análisis de Datos ---
    st.header("📊 Análisis de Datos - Visualización Interactiva")

    try:
        # Cargar datos del archivo Excel
        with st.spinner("Cargando datos del archivo Excel..."):
            df_analisis = crear_datos_desde_excel(BALANCE_PATH, BALANCE_SHEET_MUESTREO)

        if df_analisis.empty:
            st.warning("No se encontraron datos válidos en la hoja MUESTREO.")
        else:
            st.success(f"Datos cargados correctamente: {len(df_analisis)} registros válidos.")

            # Mostrar vista previa
            with st.expander("📋 Vista previa de los datos cargados"):
                st.dataframe(df_analisis, use_container_width=True) # Antes df_analisis.head(20)

            # Filtros de análisis
            st.subheader("🔍 Filtros de Análisis")

            col_f1, col_f2 = st.columns(2)
            with col_f1:
                # Filtro por evaluado
                evaluados_analisis = sorted(
                    [str(x) for x in df_analisis['EVALUADO'].unique()
                     if str(x).strip() != '' and str(x).strip().lower() != 'sin especificar']
                )
                evaluado_filtro = st.selectbox(
                    "Seleccionar EVALUADO",
                    options=["TODOS"] + evaluados_analisis,
                    index=0,
                    key="filtro_evaluado"
                )

            with col_f2:
                # Filtro por fecha
                if df_analisis['FECHA'].dtype == 'datetime64[ns]':
                    fechas_unicas = sorted(df_analisis['FECHA'].dt.date.unique())
                    fecha_filtro = st.selectbox(
                        "Filtrar por Fecha",
                        options=["TODAS"] + [str(f) for f in fechas_unicas],
                        index=0,
                        key="filtro_fecha"
                    )
                else:
                    fechas_unicas = sorted(df_analisis['FECHA'].unique())
                    fecha_filtro = st.selectbox(
                        "Filtrar por Fecha",
                        options=["TODAS"] + [str(f) for f in fechas_unicas],
                        index=0,
                        key="filtro_fecha_str"
                    )

            # Aplicar filtros
            df_filtrado = df_analisis.copy()
            if evaluado_filtro != "TODOS":
                df_filtrado = df_filtrado[df_filtrado['EVALUADO'] == evaluado_filtro]

            if fecha_filtro != "TODAS":
                if df_filtrado['FECHA'].dtype == 'datetime64[ns]':
                    df_filtrado = df_filtrado[df_filtrado['FECHA'].dt.date == pd.to_datetime(fecha_filtro).date()]
                else:
                    df_filtrado = df_filtrado[df_filtrado['FECHA'].astype(str) == fecha_filtro]

            if df_filtrado.empty:
                st.warning("No hay datos con los filtros seleccionados.")
            else:
                # Resumen estadístico
                st.subheader("📈 Resumen Estadístico")
                generar_resumen_estadistico_streamlit(df_filtrado)

                st.markdown("---")

                # Gráficos interactivos
                st.subheader("📊 Gráficos Interactivos")

                # Gráfico 1: Distribución por fecha y proceso
                st.markdown("### 1. Distribución por Fecha y Proceso")
                fig1 = crear_grafico_barras_fecha_proceso_interactivo(df_filtrado)
                if fig1:
                    st.plotly_chart(fig1, use_container_width=True, config={'displayModeBar': True})
                else:
                    st.info("No hay datos suficientes para generar este gráfico.")

                # Gráfico 2: Procesos (Pie) y Subprocesos (Barras)
                st.markdown("### 2. Análisis de Procesos y Subprocesos")

                if fecha_filtro == "TODAS" and len(df_filtrado['FECHA'].unique()) > 1:
                    # Si hay múltiples fechas, mostrar por fecha seleccionada
                    fechas_disponibles = sorted(
                        df_filtrado['FECHA_STR'].unique() if 'FECHA_STR' in df_filtrado.columns else df_filtrado[
                            'FECHA'].unique())
                    fecha_grafico = st.selectbox(
                        "Seleccionar fecha para análisis detallado:",
                        options=fechas_disponibles,
                        key="fecha_detalle"
                    )
                else:
                    fecha_grafico = fecha_filtro if fecha_filtro != "TODAS" else None

                col_g1, col_g2 = st.columns(2)
                with col_g1:
                    fig2 = crear_grafico_proceso_pie_interactivo(df_filtrado, fecha_grafico)
                    if fig2:
                        st.plotly_chart(fig2, use_container_width=True, config={'displayModeBar': True})

                with col_g2:
                    fig3 = crear_grafico_subprocesos_interactivo(df_filtrado, fecha_grafico, top_n=10)
                    if fig3:
                        st.plotly_chart(fig3, use_container_width=True, config={'displayModeBar': True})

                # Gráfico 3: Proceso vs Acciones
                st.markdown("### 3. Distribución por Proceso-Acciones")
                fig4 = crear_grafico_proceso_acciones_interactivo(df_filtrado)
                if fig4:
                    st.plotly_chart(fig4, use_container_width=True, config={'displayModeBar': True})

                # Información adicional
                st.markdown("---")
                st.info("""
                **💡 Características de los gráficos interactivos:**
                - **Hover**: Pasa el mouse sobre cualquier elemento para ver información detallada
                - **Zoom**: Haz clic y arrastra para hacer zoom en áreas específicas
                - **Pan**: Usa el botón de mano para moverte por el gráfico
                - **Descarga**: Haz clic en el icono de cámara para descargar la imagen
                - **Fullscreen**: Haz clic en el icono de pantalla completa para ver en tamaño completo
                """)

    except Exception as e:
        st.error(f"Error al cargar o analizar los datos: {str(e)}")
        st.info("""
        **Posibles soluciones:**
        1. Asegúrate de que el archivo Excel esté cerrado en Excel
        2. Verifica que la hoja 'MUESTREO' exista en el archivo
        3. Revisa que los datos tengan la estructura esperada (encabezados en fila 3)
        """)